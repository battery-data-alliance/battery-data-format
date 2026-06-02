# src/bdf/data_sources/arbin.py
"""
Arbin MITS Pro exports.

Two export variants are seen in the wild (see arbin_sample_files.md):

Format A — CSV export, space-separated column names::

    Data Point,Date Time,Test Time (s),Step Time (s),Cycle Index,Step Index,
    TC_Counter1..4,Current (A),Voltage (V),Power (W),Charge Capacity (Ah),
    Discharge Capacity (Ah),Charge Energy (Wh),Discharge Energy (Wh),
    Capacity (Ah),mAh/g,ACR (Ohm),dV/dt (V/s),Internal Resistance (Ohm),
    dQ/dV (Ah/V),dV/dQ (V/Ah),Aux_Temperature_1 (C),Aux_dT/dt_1 (C/s)

Format B — XLSX export, underscore-separated column names::

    Date_Time,Test_Time(s),Step_Time(s),Step_Index,Cycle_Index,Voltage(V),
    Current(A),Charge_Capacity(Ah),Discharge_Capacity(Ah),Charge_Energy(Wh),
    Discharge_Energy(Wh),Internal Resistance(Ohm),dV/dt(V/s),
    Aux_Voltage(V)_1..5,Aux_Temperature(°)_1..3

Both come from Arbin MITS Pro; the unit-bearing differential columns
(``dV/dt``, ``dQ/dV``, ``dV/dQ``, ``mAh/g``, ``ACR``) and the auxiliary
voltage channels have no BDF canonical equivalent and are passed through
unchanged.

Per the BDF step-quantity convention, Arbin ``Step_Index`` is the schedule
step identifier and maps to ``Step ID`` (not the derived ``Step Index / 1``).
"""
from __future__ import annotations

import re
import warnings

import numpy as np
import pandas as pd

from .base_delimited import DelimitedTextPlugin
from .excel_xlsx import ExcelXlsx

# Shared vendor-header -> BDF canonical mapping. Units are stripped by the
# normalizer before matching, so both spaced ("Charge Capacity (Ah)") and
# underscored ("Charge_Capacity(Ah)") spellings reduce to the same base slug.
_ARBIN_SYNONYMS: dict[str, list[str]] = {
    # --- Required ---
    "Test Time / s": ["test time", "test_time"],
    "Voltage / V":   ["voltage"],
    "Current / A":   ["current"],

    # --- Recommended ---
    "Unix Time / s":   ["date time", "date_time"],
    "Cycle Count / 1": ["cycle index", "cycle_index"],
    "Step ID":         ["step index", "step_index"],
    "Step Time / s":   ["step time", "step_time"],

    # --- Optional (capacities / energies) ---
    "Charging Capacity / Ah":    ["charge capacity", "charge_capacity"],
    "Discharging Capacity / Ah": ["discharge capacity", "discharge_capacity"],
    "Charging Energy / Wh":      ["charge energy", "charge_energy"],
    "Discharging Energy / Wh":   ["discharge energy", "discharge_energy"],
    "Power / W":                 ["power"],

    # ACR is AC resistance; Arbin also exports a separate DC "Internal
    # Resistance" column which is the one BDF treats as internal_resistance.
    "Internal Resistance / ohm": ["internal resistance", "internal_resistance"],

    # --- Optional (auxiliary surface temperatures, channels 1..5) ---
    "Surface Temperature T1 / degC": ["aux_temperature_1", "aux_temperature 1"],
    "Surface Temperature T2 / degC": ["aux_temperature_2", "aux_temperature 2"],
    "Surface Temperature T3 / degC": ["aux_temperature_3", "aux_temperature 3"],
    "Surface Temperature T4 / degC": ["aux_temperature_4", "aux_temperature 4"],
    "Surface Temperature T5 / degC": ["aux_temperature_5", "aux_temperature 5"],
}

# Format A (CSV) signature: columns unique to the Arbin MITS Pro CSV export.
# Kept tight so it does not out-score other comma-delimited cyclers (Neware,
# LANDT, Novonix) that also carry generic charge/discharge capacity columns.
_ARBIN_CSV_TOKENS = (
    r"\bdata\s*point\b",
    r"\btc[_\s]*counter\d",
    r"\baux[_\s]*temperature",
)

# Format B (XLSX) signature: the underscore-joined Arbin header spellings,
# distinctive enough to separate an Arbin workbook from a generic spreadsheet.
_ARBIN_XLSX_TOKENS = (
    r"\btest_time\b",
    r"\bstep_index\b",
    r"\baux_temperature\b",
    r"\baux_voltage\b",
)

# Directional accumulators that BDF defines as never resetting between steps or
# cycles. Arbin exports them as test-wide accumulators but the test schedule may
# issue an occasional counter reset (the column drops to ~0 at a step boundary).
# These columns are re-accumulated through such resets so the output matches the
# BDF definition (see CHANGELOG / ontology charging_capacity_ah definition).
_ARBIN_ACCUMULATORS = (
    "Charging Capacity / Ah",
    "Discharging Capacity / Ah",
    "Charging Energy / Wh",
    "Discharging Energy / Wh",
)


def _reaccumulate_resets(s: pd.Series) -> tuple[pd.Series, bool]:
    """Re-integrate an accumulator through schedule-driven counter resets.

    A reset is a clear downward jump to (near) baseline — distinct from float
    noise, which is clamped to a zero increment. Returns (series, changed). When
    no reset is detected the original series is returned untouched so already-
    monotonic instrument values are preserved bit-for-bit.
    """
    v = pd.to_numeric(s, errors="coerce").to_numpy(dtype="float64")
    if v.size < 2 or not np.isfinite(v).any():
        return s, False

    prev, cur = v[:-1], v[1:]
    delta = cur - prev
    # Genuine reset: a real drop that lands near baseline (well below the
    # pre-reset level), not a tiny numerical wobble at a high running value.
    is_reset = (delta < 0) & (cur < 0.5 * prev) & ((prev - cur) > 1e-6)
    if not is_reset.any():
        return s, False

    # On a reset row the true increment is the post-reset value (accumulation
    # restarting from zero); other decreases are noise -> zero increment.
    corrected = np.where(is_reset, np.maximum(cur, 0.0), np.where(delta < 0, 0.0, delta))
    out = np.empty_like(v)
    out[0] = v[0] if np.isfinite(v[0]) else 0.0
    out[1:] = out[0] + np.cumsum(corrected)
    return pd.Series(out, index=s.index), True


def _fixup_arbin_accumulators(df: pd.DataFrame) -> pd.DataFrame:
    out = df
    changed: list[str] = []
    for col in _ARBIN_ACCUMULATORS:
        if col not in df.columns:
            continue
        new, did = _reaccumulate_resets(df[col])
        if did:
            if out is df:
                out = df.copy()
            out[col] = new
            changed.append(col)
    if changed:
        warnings.warn(
            "Arbin: re-accumulated "
            + ", ".join(changed)
            + " across instrument counter reset(s) so the directional "
            "charge/discharge accumulators never reset between steps or cycles, "
            "per the BDF definition.",
            UserWarning,
            stacklevel=2,
        )
    return out


class ArbinCSV(DelimitedTextPlugin):
    """Arbin MITS Pro CSV export (Format A)."""

    id = "arbin-csv"
    exts = (".csv",)
    default_encoding = "utf-8-sig"
    ragged_row_policy = "fold_last"

    # Arbin timestamps look like "10/10/2023 16:09:14.661" (MM/DD/YYYY).
    timestamp_candidate_patterns = (r"^date[_\s-]*time$",)
    assume_naive_tz = "UTC"

    header_token_patterns = _ARBIN_CSV_TOKENS
    column_synonyms = _ARBIN_SYNONYMS

    def fixup(self, df: pd.DataFrame) -> pd.DataFrame:
        return _fixup_arbin_accumulators(super().fixup(df))


class ArbinExcel(ExcelXlsx):
    """Arbin MITS Pro XLSX export (Format B)."""

    id = "arbin-xlsx"
    exts = (".xlsx", ".xlsm", ".xls")
    column_synonyms = _ARBIN_SYNONYMS

    # Normalise the degree-symbol auxiliary temperature headers
    # ("Aux_Temperature(°)_1") to a clean, unit-bearing form the normalizer
    # parses reliably ("Aux_Temperature_1 (degC)").
    _AUX_TEMP_RE = re.compile(
        r"^aux[_\s]*temperature\s*\(\s*[°°c]*\s*\)\s*_?(\d+)$",
        re.IGNORECASE,
    )

    # Lower-cased tokens that mark the time-series sheet inside a multi-sheet
    # Arbin workbook (the data lives on a "Channel_N" sheet; sheet 0 is often a
    # "Global_Info" / "TEST REPORT" metadata sheet).
    _DATA_SHEET_REQUIRED = ("test_time", "voltage", "current")

    def sniff(self, path, head):
        result = super().sniff(path, head)
        # Confirm it is specifically an Arbin workbook by peeking the header row
        # of the data sheet (any sheet, not just the first/active one).
        sheet, headers = self._find_data_sheet(path)
        header_text = " ".join(headers)
        if sheet is not None and any(re.search(p, header_text, re.I) for p in _ARBIN_XLSX_TOKENS):
            result.confidence = min(result.confidence + 0.25, 1.0)
            result.reason = (result.reason + "+arbin").strip("+")
            result.meta["arbin"] = True
            result.meta["sheet"] = sheet
        else:
            # Defer to the generic excel-xlsx reader for non-Arbin workbooks.
            result.confidence = min(result.confidence, 0.30)
        return result

    @classmethod
    def _find_data_sheet(cls, path) -> tuple[str | None, list[str]]:
        """Return (sheet_name, header) for the Arbin time-series sheet.

        Scans every sheet's first row and prefers the one carrying the required
        Arbin data columns. Falls back to the active sheet, then the first sheet,
        so single-sheet workbooks keep working.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                first_sheet: tuple[str, list[str]] | None = None
                active_name = getattr(wb.active, "title", None)
                active_sheet: tuple[str, list[str]] | None = None
                for ws in wb.worksheets:
                    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                    headers = [str(c) for c in first if c is not None]
                    low = " ".join(headers).lower()
                    if all(tok in low for tok in cls._DATA_SHEET_REQUIRED):
                        return ws.title, headers
                    if first_sheet is None:
                        first_sheet = (ws.title, headers)
                    if ws.title == active_name:
                        active_sheet = (ws.title, headers)
                return active_sheet or first_sheet or (None, [])
            finally:
                wb.close()
        except Exception:
            return None, []

    def parse(self, path) -> pd.DataFrame:
        sheet, _ = self._find_data_sheet(path)
        if sheet is not None:
            try:
                df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
                df.columns = [str(c).strip().lstrip("﻿") for c in df.columns]
            except Exception:
                df = super().parse(path)
        else:
            df = super().parse(path)

        renames: dict[str, str] = {}
        for col in df.columns:
            m = self._AUX_TEMP_RE.match(str(col).strip())
            if m:
                renames[col] = f"Aux_Temperature_{m.group(1)} (degC)"
        if renames:
            df = df.rename(columns=renames)
        return df

    def fixup(self, df: pd.DataFrame) -> pd.DataFrame:
        return _fixup_arbin_accumulators(super().fixup(df))
